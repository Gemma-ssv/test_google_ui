import time
import openpyxl
import random
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



def create_excel_file(filename):
    wb = Workbook()
    ws = wb.active

    headers = ['email', 'password', 'name', 'surname', 'date_of_birth', 'backup_email']
    ws.append(headers)

    wb.save(filename)
    print(f"Файл {filename} успешно создан.")
    return filename

def change_email_password(email, password, new_password):
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    # Инициализация драйвера
    with webdriver.Chrome(options=options) as driver:
        driver.get('https://www.google.com/intl/ru/gmail/about/')

        # Небольшая задержка для имитации действий человека
        time.sleep(random.uniform(1, 3))

        # Находим кнопку и кликаем на неё
        buttons = driver.find_elements(By.CLASS_NAME, "button")
        buttons[1].click()

        # Задержка перед вводом email
        time.sleep(random.uniform(2, 4))

        # Ввод email
        email_field = driver.find_element(By.ID, 'identifierId')
        for char in email:
            email_field.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

        # Задержка перед нажатием кнопки "Далее"
        time.sleep(random.uniform(2, 4))

        # Находим и кликаем кнопку "Далее"
        next_button = driver.find_element(By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ VfPpkd-LgbsSe-OWXEXe-dgl2Hf nCP5yc AjY5Oe DuMIQc LQeN7 BqKGqe Jskylb TrZEUc lw1w4b"]')
        next_button.click()

        # Задержка перед вводом пароля
        time.sleep(random.uniform(3, 5))

        # Ввод пароля
        password_field = driver.find_element(By.NAME, 'Passwd')
        for char in password:
            password_field.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

        # Задержка перед нажатием кнопки "Далее"
        time.sleep(random.uniform(2, 4))

        # Находим и кликаем кнопку "Далее"
        password_next_button = driver.find_elements(By.TAG_NAME, 'button')[1]
        password_next_button.click()

        # Задержка перед проверкой наличия сообщения "Не сейчас"
        time.sleep(random.uniform(3, 5))

        # Проверка и клик на "Не сейчас", если сообщение присутствует
        if driver.find_elements(By.TAG_NAME, 'span') == 'Не сейчас':
            not_now_button = driver.find_element(By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-dgl2Hf ksBjEc lKxP2d LQeN7 BqKGqe eR0mzb TrZEUc lw1w4b"]')
            not_now_button.click()
            time.sleep(random.uniform(3, 5))

        # Переход по ссылкам с имитацией действий человека
        links = [
            ('//a[@class="FH"]', 3),
            ('//button[@class="Tj"]', 3),
            ('(//a[@class="f0 LJOhwe"])[2]', 3),
            ('//a[text()="Изменить пароль"]', 3)
        ]

        for xpath, delay in links:
            link = driver.find_element(By.XPATH, xpath)
            actions = ActionChains(driver)
            actions.move_to_element(link).perform()
            time.sleep(random.uniform(1, 2))
            link.click()
            time.sleep(delay)
            # Переключение на новую вкладку, если она открылась
            if len(driver.window_handles) > 1:
                new_window = driver.window_handles[-1]  # Последняя открытая вкладка
                driver.switch_to.window(new_window)
                time.sleep(random.uniform(3, 5))
                
                # Ожидание, пока элементы станут видимыми
                input1 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'i6')))
                
                for char in new_password:
                    input1.send_keys(char)
                    time.sleep(random.uniform(0.1, 0.3))
                time.sleep(random.uniform(1, 5))
                
                input2 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'i12')))
                for char in new_password:
                    input2.send_keys(char)
                    time.sleep(random.uniform(0.1, 0.3))
                time.sleep(random.uniform(3, 5))
                
                # Ожидание, пока кнопка станет видимой и кликабельной
                change_password_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//html/body/c-wiz/div/div[2]/div[2]/c-wiz/div/div[4]/form/div/div[2]/div/div/button[@jsname="Pr7Yme"]')))
                
                # Прокручиваем страницу до кнопки
                driver.execute_script("arguments[0].scrollIntoView(true);", change_password_button)
                time.sleep(random.uniform(2, 5))  
                # Находим кнопку и имитируем клик мышью
                actions = ActionChains(driver)
                actions.move_to_element(change_password_button).click().perform()
                time.sleep(random.uniform(4, 5))
                
                # Ожидание, пока модальное окно станет видимым
                WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//div[@aria-modal="true"]')))

                # Ожидание, пока кнопка "Сменить пароль" станет видимой и кликабельной
                change_password_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//button[@data-mdc-dialog-action="ok"]'))
                )

                # Прокручиваем страницу до кнопки
                driver.execute_script("arguments[0].scrollIntoView(true);", change_password_button)
                time.sleep(random.uniform(2, 5))

                # Находим кнопку и имитируем клик мышью
                actions = ActionChains(driver)
                actions.move_to_element(change_password_button).click().perform()
                time.sleep(random.uniform(3, 5))
                
                return new_password 

def change_name_and_surname_to_email(email, password, new_name, new_surname):
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    # Инициализация драйвера
    with webdriver.Chrome(options=options) as driver:
        driver.get('https://www.google.com/intl/ru/gmail/about/')

        # Небольшая задержка для имитации действий человека
        time.sleep(random.uniform(1, 3))

        # Находим кнопку и кликаем на неё
        buttons = driver.find_elements(By.CLASS_NAME, "button")
        buttons[1].click()

        # Задержка перед вводом email
        time.sleep(random.uniform(2, 4))

        # Ввод email
        email_field = driver.find_element(By.ID, 'identifierId')
        for char in email:
            email_field.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

        # Задержка перед нажатием кнопки "Далее"
        time.sleep(random.uniform(2, 4))

        # Находим и кликаем кнопку "Далее"
        next_button = driver.find_element(By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ VfPpkd-LgbsSe-OWXEXe-dgl2Hf nCP5yc AjY5Oe DuMIQc LQeN7 BqKGqe Jskylb TrZEUc lw1w4b"]')
        next_button.click()

        # Задержка перед вводом пароля
        time.sleep(random.uniform(3, 5))

        # Ввод пароля
        password_field = driver.find_element(By.NAME, 'Passwd')
        for char in password:
            password_field.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

        # Задержка перед нажатием кнопки "Далее"
        time.sleep(random.uniform(2, 4))

        # Находим и кликаем кнопку "Далее"
        password_next_button = driver.find_elements(By.TAG_NAME, 'button')[1]
        password_next_button.click()

        # Задержка перед проверкой наличия сообщения "Не сейчас"
        time.sleep(random.uniform(3, 5))

        # Проверка и клик на "Не сейчас", если сообщение присутствует
        if driver.find_elements(By.TAG_NAME, 'span') == 'Не сейчас':
            not_now_button = driver.find_element(By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-dgl2Hf ksBjEc lKxP2d LQeN7 BqKGqe eR0mzb TrZEUc lw1w4b"]')
            not_now_button.click()
            time.sleep(random.uniform(3, 5))

        # Переход по ссылкам с имитацией действий человека
        links = [
            ('//a[@class="FH"]', 3),
            ('//button[@class="Tj"]', 3),
            ('(//a[@class="f0 LJOhwe"])[2]', 3),
            ('//a[text()="Другие настройки аккаунта Google"]', 3)
        ]

        for xpath, delay in links:
            link = driver.find_element(By.XPATH, xpath)
            actions = ActionChains(driver)
            actions.move_to_element(link).perform()
            time.sleep(random.uniform(1, 2))
            link.click()
            time.sleep(delay)
            # Переключение на новую вкладку, если она открылась
            if len(driver.window_handles) > 1:
                new_window = driver.window_handles[-1]  # Последняя открытая вкладка
                driver.switch_to.window(new_window)
                time.sleep(random.uniform(3, 5))
                driver.find_element(By.XPATH, '//a[@data-rid="10003"]').click()
                time.sleep(random.uniform(3, 5))
                
                # Ожидание, пока элемент станет видимым и кликабельным
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//a[@data-rid="10090"]'))).click()
                time.sleep(random.uniform(4, 5))
                
                change_name_button = driver.find_elements(By.XPATH, '//a[@jsname="hSRGPd"]')
                change_name_button[2].click()
                time.sleep(random.uniform(3, 5))
                
                # Находим элементы поля ввода
                name = driver.find_element(By.ID, 'i7')
                surname = driver.find_element(By.ID, 'i12')

                # Проверяем, заполнено ли поле name, и если да, очищаем его
                if name.get_attribute('value'):
                    name.clear()
                for char in new_name:
                    name.send_keys(char)
                    time.sleep(random.uniform(0.1, 0.3))
                time.sleep(random.uniform(1, 5))

                # Проверяем, заполнено ли поле surname, и если да, очищаем его
                if surname.get_attribute('value'):
                    surname.clear()
                for char in new_surname:
                    surname.send_keys(char)
                    time.sleep(random.uniform(0.1, 0.3))
                time.sleep(random.uniform(1, 5))
                
                # Ожидание, пока кнопка станет видимой и кликабельной
                change_name_and_surname_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="yDmH0d"]/c-wiz/div/div[2]/div[2]/c-wiz/div[2]/div/div/div[3]/div[2]/div/div/button')))
                
                # Прокручиваем страницу до кнопки
                driver.execute_script("arguments[0].scrollIntoView(true);", change_name_and_surname_button)
                time.sleep(random.uniform(2, 5))  
                # Находим кнопку и имитируем клик мышью
                actions = ActionChains(driver)
                actions.move_to_element(change_name_and_surname_button).click().perform()
                
                ret_list = [new_name, surname]
                
                return ret_list

def get_data_email(email, password):
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    # Инициализация драйвера
    with webdriver.Chrome(options=options) as driver:
        driver.get('https://www.google.com/intl/ru/gmail/about/')

        # Небольшая задержка для имитации действий человека
        time.sleep(random.uniform(1, 3))

        # Находим кнопку и кликаем на неё
        buttons = driver.find_elements(By.CLASS_NAME, "button")
        buttons[1].click()

        # Задержка перед вводом email
        time.sleep(random.uniform(2, 4))

        # Ввод email
        email_field = driver.find_element(By.ID, 'identifierId')
        for char in email:
            email_field.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

        # Задержка перед нажатием кнопки "Далее"
        time.sleep(random.uniform(2, 4))

        # Находим и кликаем кнопку "Далее"
        next_button = driver.find_element(By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-k8QpJ VfPpkd-LgbsSe-OWXEXe-dgl2Hf nCP5yc AjY5Oe DuMIQc LQeN7 BqKGqe Jskylb TrZEUc lw1w4b"]')
        next_button.click()

        # Задержка перед вводом пароля
        time.sleep(random.uniform(3, 5))

        # Ввод пароля
        password_field = driver.find_element(By.NAME, 'Passwd')
        for char in password:
            password_field.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))

        # Задержка перед нажатием кнопки "Далее"
        time.sleep(random.uniform(2, 4))

        # Находим и кликаем кнопку "Далее"
        password_next_button = driver.find_elements(By.TAG_NAME, 'button')[1]
        password_next_button.click()

        # Задержка перед проверкой наличия сообщения "Не сейчас"
        time.sleep(random.uniform(3, 5))

        # Проверка и клик на "Не сейчас", если сообщение присутствует
        if driver.find_elements(By.TAG_NAME, 'span') == 'Не сейчас':
            not_now_button = driver.find_element(By.XPATH, '//button[@class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-dgl2Hf ksBjEc lKxP2d LQeN7 BqKGqe eR0mzb TrZEUc lw1w4b"]')
            not_now_button.click()
            time.sleep(random.uniform(3, 5))

        # Переход по ссылкам с имитацией действий человека
        links = [
            ('//a[@class="FH"]', 3),
            ('//button[@class="Tj"]', 3),
            ('(//a[@class="f0 LJOhwe"])[2]', 3),
            ('//a[text()="Другие настройки аккаунта Google"]', 3)
        ]

        for xpath, delay in links:
            link = driver.find_element(By.XPATH, xpath)
            actions = ActionChains(driver)
            actions.move_to_element(link).perform()
            time.sleep(random.uniform(1, 2))
            link.click()
            time.sleep(delay)
            # Переключение на новую вкладку, если она открылась
            if len(driver.window_handles) > 1:
                new_window = driver.window_handles[-1]  # Последняя открытая вкладка
                driver.switch_to.window(new_window)
                time.sleep(random.uniform(3, 5))
                driver.find_element(By.XPATH, '//a[@data-rid="10003"]').click()
                time.sleep(random.uniform(3, 5))
                datas = driver.find_elements(By.CLASS_NAME, 'bJCr1d')

                name_surname = datas[3].text.split(' ')
                name = name_surname[0]
                surname = name_surname[1]
                
                date_of_birth = datas[4].text
                try:               
                    email = datas[6].text
                    backup_email = datas[7].text
                except Exception:
                    email = datas[6].text
                    backup_email = 'None'
                passwd = password        
    data_list = [email, passwd, name, surname, date_of_birth, backup_email]
    return data_list
               
def save_data_email(filename, data):
    try:
        # Попытка открыть существующий файл
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # Если файл не найден, создаем новый
        wb = Workbook()
        ws = wb.active
        headers = ['email', 'password', 'name', 'surname', 'date_of_birth', 'backup_email']
        ws.append(headers)

    # Добавляем новую строку с данными
    ws.append([data[0], data[1], data[2], data[3], data[4], data[5]])

    # Сохраняем файл
    wb.save(filename)
    print(f"Данные успешно добавлены в файл {filename}.")
    
    
# ВАРИАНТ ИСПОЛЬЗОВАНИЯ

# СОЗДАЁМ ФАЙЛ
file = create_excel_file('test.xlsx')

# ОБНОВЛЯЕМ ПАРОЛЬ
new_passw = change_email_password('test@gmail.com','test_password', 'new_test_password')

# ОБНОВЛЯЕМ ИМЯ И ФАМИЛИЯ
change_name_and_surname_to_email('test@gmail.com', new_passw,'ТЕСТ', 'ТЕСТОВ')

# ПОЛУЧАЕМ ОБНОВЛЕННЫЕ ДАННЫЕ
data_list = get_data_email('test@gmail.com', new_passw)

# СОХРАНЯЕМ ДАННЫЕ В ФАЙЛ
save_data_email(file, data_list)

